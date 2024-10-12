import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Función para crear o cargar el archivo Excel
def crear_o_cargar_archivo(nombre_archivo):
    try:
        libro = openpyxl.load_workbook(nombre_archivo)
        print(f"Archivo '{nombre_archivo}' cargado exitosamente.")
    except FileNotFoundError:
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Gastos"
        hoja.append(["Fecha", "Descripción", "Monto"])  # Cabeceras
        print(f"Archivo '{nombre_archivo}' creado con una nueva hoja de gastos.")
    return libro

# Función para ingresar gastos
def ingresar_gastos():
    gastos = []
    while True:
        fecha = input("Ingresa la fecha (YYYY-MM-DD): ")
        descripcion = input("Ingresa la descripción del gasto: ")
        while True:
            try:
                monto = float(input("Ingresa el monto del gasto: "))
                break
            except ValueError:
                print("Por favor, ingresa un número válido para el monto.")
        
        gastos.append((fecha, descripcion, monto))
        
        continuar = input("¿Deseas agregar otro gasto? (s/n): ").lower()
        if continuar != 's':
            break
    return gastos

# Función para guardar los gastos en el archivo Excel
def guardar_en_excel(gastos, libro, nombre_archivo):
    hoja = libro["Gastos"]
    for gasto in gastos:
        hoja.append(gasto)
    libro.save(nombre_archivo)
    print(f"Gastos guardados en el archivo '{nombre_archivo}'.")

# Función para generar el informe de resumen
def generar_resumen(gastos):
    if not gastos:
        print("No se ingresaron gastos.")
        return
    
    total_gastos = len(gastos)
    gasto_mayor = max(gastos, key=lambda x: x[2])
    gasto_menor = min(gastos, key=lambda x: x[2])
    monto_total = sum(gasto[2] for gasto in gastos)
    
    print("\nResumen de Gastos:")
    print(f"Total de gastos: {total_gastos}")
    print(f"Gasto más caro: {gasto_mayor[1]} el {gasto_mayor[0]} por {gasto_mayor[2]:.2f}")
    print(f"Gasto más barato: {gasto_menor[1]} el {gasto_menor[0]} por {gasto_menor[2]:.2f}")
    print(f"Monto total de gastos: {monto_total:.2f}")
    
    return total_gastos, gasto_mayor, gasto_menor, monto_total

# Función para guardar el informe en Excel
def guardar_resumen_en_excel(libro, total_gastos, gasto_mayor, gasto_menor, monto_total, nombre_archivo):
    hoja = libro["Gastos"]
    ultima_fila = hoja.max_row + 2
    hoja[f"A{ultima_fila}"] = "Resumen de Gastos"
    hoja[f"A{ultima_fila+1}"] = f"Total de gastos: {total_gastos}"
    hoja[f"A{ultima_fila+2}"] = f"Gasto más caro: {gasto_mayor[1]} el {gasto_mayor[0]} por {gasto_mayor[2]:.2f}"
    hoja[f"A{ultima_fila+3}"] = f"Gasto más barato: {gasto_menor[1]} el {gasto_menor[0]} por {gasto_menor[2]:.2f}"
    hoja[f"A{ultima_fila+4}"] = f"Monto total de gastos: {monto_total:.2f}"
    
    libro.save(nombre_archivo)
    print(f"Resumen guardado en el archivo '{nombre_archivo}'.")

# Función principal
def main():
    nombre_archivo = "informe_gastos.xlsx"
    libro = crear_o_cargar_archivo(nombre_archivo)
    
    # Ingreso de datos de gastos
    gastos = ingresar_gastos()
    
    # Guardar los gastos en el archivo Excel
    guardar_en_excel(gastos, libro, nombre_archivo)
    
    # Generar y mostrar el resumen de gastos
    total_gastos, gasto_mayor, gasto_menor, monto_total = generar_resumen(gastos)
    
    # Guardar el resumen en el archivo Excel
    guardar_resumen_en_excel(libro, total_gastos, gasto_mayor, gasto_menor, monto_total, nombre_archivo)

if __name__ == "__main__":
    main()
